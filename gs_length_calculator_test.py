import cv2
import sys
import os, csv
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import glob
from openpyxl import load_workbook
import json
#filepath = '/app2/Joint_Data/no_pad_train/img_mask/'
filepath = '/app/pilot/img_mask/'

def save_to_excel(list_info, order):
    temp_array = []
    #list_info.append(labels)
    #print(len(list_info))
    print(len(list_info))
    for idx, v in enumerate(list_info):
        # print(len(list_info[idx]))
        temp_array.append((
                          list_info[idx][0], list_info[idx][1], list_info[idx][2], list_info[idx][3], list_info[idx][4],
                          list_info[idx][5], list_info[idx][6], list_info[idx][7]))
        # temp_array.append((list_info[idx][0], list_info[idx][1], list_info[idx][2], list_info[idx][3], list_info[idx][4], list_info[idx][5]))
    # 0308_gs_minlength
    pd.DataFrame(temp_array).to_excel('./0421_oai_total' + order + '.xlsx', index=False)
def save_to_excel_test_Threshold(list_info, order, Threshold):
    if order == 'train':
        filename = './0621_oai_train_set_' + str(Threshold) + '.xlsx'
    elif order == 'test':
        # filename = './0705_oai_test_set_0_4_set_25.xlsx'
        filename = './0706_Threshold_test/0706_oai_test_set_' + str(Threshold) + '.xlsx'
    else:
        # filename = './0618_pilot_set_0_2_set.xlsx'
        #filename = './pilot_excel/0714_pilot_set_' + str(Threshold) + '.xlsx'
        filename = '/app/threshold_test/0824_pilot_set_' + str(Threshold) + '.xlsx'

    temp_array = []
    #list_info.append(labels)
    #print(len(list_info))
    print(len(list_info))
    for idx, v in enumerate(list_info):
        # print(len(list_info[idx]))
        # temp_array.append((list_info[idx][0], list_info[idx][1], list_info[idx][2], list_info[idx][3], list_info[idx][4], list_info[idx][5], list_info[idx][6], list_info[idx][7]))
        # add gs list
        temp_array.append((
                          list_info[idx][0], list_info[idx][1], list_info[idx][2], list_info[idx][3], list_info[idx][4],
                          list_info[idx][5], list_info[idx][6], list_info[idx][7], list_info[idx][8], list_info[idx][9],
                          list_info[idx][10], list_info[idx][11], list_info[idx][12], list_info[idx][13],
                          list_info[idx][14], list_info[idx][15],list_info[idx][16], list_info[idx][17]))

    # 0308_gs_minlength
    pd.DataFrame(temp_array).to_excel(filename, index=False)
def save_to_excel_test(list_info, order):
    if order == 'train':
        filename = './0621_oai_train_set_0_2_set.xlsx'
    elif order == 'test':
        filename = './0621_oai_test_set_0_2_set.xlsx'
    else:
        #filename = './0618_pilot_set_0_2_set.xlsx'
        filename = '/app/0824_pilot_set_0_2_set.xlsx'

    temp_array = []
    #list_info.append(labels)
    #print(len(list_info))
    print(len(list_info))
    for idx, v in enumerate(list_info):
        # print(len(list_info[idx]))
        temp_array.append((
                          list_info[idx][0], list_info[idx][1], list_info[idx][2], list_info[idx][3], list_info[idx][4], list_info[idx][5], list_info[idx][6], list_info[idx][7]))
        # temp_array.append((list_info[idx][0], list_info[idx][1], list_info[idx][2], list_info[idx][3], list_info[idx][4], list_info[idx][5]))
    # 0308_gs_minlength
    pd.DataFrame(temp_array).to_excel(filename, index=False)
# seg_path = '/app2/Joint_Data/0205_half_learning_intersection_no_pad_no_rotation_tf2/new_result_train_set/y_pred/positive'
#seg_path = '/app2/Joint_Data/0216_half_learning_intersection_no_pad_tf2_batch4/new_result_train/y_pred/positive'

seg_path = '/app/0802_new_img_half_augment_data_relabeled_redraw_data/pilot_result_new/y_pred/positive'

compare_flag = False
std_multiple_list = [1, 2, 3, 4, 5]
T_multiple_list = [0.1, 0.2, 0.3, 0.4, 0.5]

def load_excel_oa_base(filename):
    workbook = load_workbook(filename)
    # sheet = workbook.active
    file_names = []
    kl_names = []
    labels_names = []
    ages = []
    genders = []
    # ws = ['0', '1', '2', '3', '4']
    ws = ['Sheet1']
    temp_list = []
    for sh_num in ws:
        sheet = workbook[sh_num]
        for c1, c2, c3, c4, c5, c6, c7, c8, c9, c10, c11, c12, c13, c14, c15, c16, c17, c18, c19, c20, c21, c22, c23, c24, c25, c26 in \
        sheet[sheet.dimensions]:
            # print("{0:8} {1:8}".format(c1.value, c6.value))
            # if c22.value == '-1':
            #     continue
            # else:
            if c2.value == '2':
                temp_name = c5.value + 'R' + c3.value + '.jpg'
            else:
                temp_name = c5.value + 'L' + c3.value + '.jpg'
            file_names.append(temp_name)
            kl_names.append(c8.value)
            labels_names.append(c22.value)
            gender = 0
            if c6.value == 'M':
                genders.append(0)
                gender = 0
            else:
                gender = 1
                genders.append(1)
            # temp_list.append((temp_name, float(c7.value) / 12.0, gender))
            ages.append(float(c7.value) / 12.0)
    return file_names, ages, genders


# test original image & keypoint
def calculator_jsw_keypoint_test(original, keypoint_list, name):
    for keypoint in keypoint_list:
        cv2.line(original, (keypoint[0], keypoint[1]), (keypoint[0], keypoint[1]), (0, 0, 255), 5)
    # cv2.line(img_jsw, (tm_xmin, 0), (tm_xmin, img_jsw.shape[1]), (255, 0, 0), 5)
    cv2.imwrite('/app/oai_total_matching_img/' + name, original)

def abnormal_jsw_test_Threshold(left_avg, right_avg, left_std, right_std, img, name, test_img, Threshold):
    standard_left_X_rate = left_avg
    standard_right_X_rate = right_avg
    std_multiple = 4
    # print(standard_left_X_rate)
    # print(standard_right_X_rate)
    # print(img.shape)
    points = json_to_str(name.replace('.jpg', '.json'))
    for point_info in points:
        if point_info[0] == 'fl':
            gs_fl = (int(point_info[1]), int(point_info[2]))
        elif point_info[0] == 'fm':
            gs_fm = (int(point_info[1]), int(point_info[2]))
        elif point_info[0] == 'tl':
            gs_tl = (int(point_info[1]), int(point_info[2]))
        elif point_info[0] == 'tm':
            gs_tm = (int(point_info[1]), int(point_info[2]))

#    mask_img = cv2.imread(os.path.join('/app2/pilot/img_mask', name.replace('.jpg', '.png')))
    mask_img = cv2.imread(os.path.join('/app/pilot/img_mask', name.replace('.jpg', '.png')))

    mask_img *= 255

    if gs_fl[0] < gs_tl[0]:
        gs_left_near_X = abs(gs_tl[0])
    else:
        gs_left_near_X = abs(gs_fl[0])
    if gs_fm[0] > gs_tm[0]:
        gs_right_near_X = abs(gs_tm[0])
    else:
        gs_right_near_X = abs(gs_fm[0])

    ret, img_jsw = cv2.threshold(img, int(Threshold * 255.), 255, 0)
    # for joint length
    ret, img_length = cv2.threshold(img, int(Threshold * 255.), 255, 0)
    # print(img_length.shape)
    img_length = cv2.cvtColor(img_length, cv2.COLOR_BGR2GRAY)

    _, contours, _ = cv2.findContours(img_length, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
    # left XY, right XY in joint object
    leftmost = contours[0][contours[0][:, :, 0].argmin()][0]
    rightmost = contours[0][contours[0][:, :, 0].argmax()][0]
    for c in contours:
        temp_left = c[c[:, :, 0].argmin()][0]
        temp_right = c[c[:, :, 0].argmax()][0]
        if leftmost[0] > temp_left[0]: leftmost = temp_left
        if rightmost[0] < temp_right[0]: rightmost = temp_right

    templrx = int(((rightmost[0] - leftmost[0]) / 2) + leftmost[0])
    # img_jsw = cv2.cvtColor(img_jsw, cv2.COLOR_GRAY2BGR)
    a1 = int((templrx - leftmost[0]) * 0.1 + leftmost[0])
    a4 = int((templrx - leftmost[0]) * 0.6 + leftmost[0])
    b1 = int((rightmost[0] - templrx) * 0.4 + templrx)
    b4 = int((rightmost[0] - templrx) * 0.9 + templrx)
    if abs(leftmost[0] - int(img_jsw.shape[1] * standard_left_X_rate)) > img_jsw.shape[1] * (
            std_multiple * left_std):
        left_X = int(img_jsw.shape[1] * standard_left_X_rate)
        # left_X = leftmost[0]
    else:
        left_X = leftmost[0]
    if abs(rightmost[0] - int(img_jsw.shape[1] * standard_right_X_rate)) > img_jsw.shape[1] * (
            std_multiple * right_std):
        right_X = int(img_jsw.shape[1] * standard_right_X_rate)
        # right_X = rightmost[0]
    else:
        right_X = rightmost[0]
    left_near_X = abs(left_X)
    right_near_X = abs(right_X)
    # left extraction & probe left joint space min, joint space min's average
    left_width_min = img_jsw.shape[0]
    left_X_list = []
    left_length_list = []
    gs_left_X_list = []
    gs_left_length_list = []
    gs_right_X_list = []
    gs_right_length_list = []
    # for x in range(left_X, int(img_jsw.shape[1] * 0.4)):
    # small_left_range = int((2 / 5) * (img_jsw.shape[1] * 0.4 - left_X))
    # cv2.line(img_jsw, (leftmost[0], 0), (leftmost[0], img_jsw.shape[1]), (255, 0, 0), 5)
    # cv2.line(img_jsw, (rightmost[0], 0), (rightmost[0], img_jsw.shape[1]), (255, 0, 0), 5)
    # cv2.imwrite('./oai_line_test/ ' + name, img_jsw)
    large_left_range = int((4 / 10) * ((rightmost[0] - leftmost[0]) * 0.4))
    for x in range(leftmost[0],
                   leftmost[0] + large_left_range):
        templeft_img = img_jsw[:, x]
        testlen = len(templeft_img[templeft_img >= 200])
        left_X_list.append(x)
        left_length_list.append(testlen)

    # left_avg = np.average(left_length_list) / (right_X - left_X)
    # left_min = np.min(left_length_list) / (right_X - left_X)
    # left_std = np.std(left_length_list) / (right_X - left_X)
    left_avg = np.average(left_length_list) / (rightmost[0] - leftmost[0])
    left_min = np.min(left_length_list) / (rightmost[0] - leftmost[0])
    left_std = np.std(left_length_list) / (rightmost[0] - leftmost[0])
    # print(left_avg)
    # # print(left_std)
    # left_temp_length_list = []
    # left_temp_X_list = []
    # for idx, left_x_length in enumerate(left_min_list):
    #     if abs(left_avg - left_x_length) <= left_std:
    #         left_temp_length_list.append(left_x_length)
    #         left_temp_X_list.append(left_min_X_list[idx])
    # # print(left_temp_length_list)
    # print('left min length :', np.average(left_temp_length_list) / abs(right_X - left_X))
    right_width_min = img_jsw.shape[0]
    right_X_list = []
    right_length_list = []
    # print(right_X)
    # print(int((4 / 5) * (right_X - img_jsw.shape[1] * 0.6)))
    # for x in range(int(img_jsw.shape[1] * 0.6), right_X):
    small_right_range = int((6 / 10) * ((rightmost[0] - leftmost[0]) * 0.4))
    large_right_range = int((4 / 5) * (rightmost[0] - img_jsw.shape[1] * 0.6))

    for x in range(leftmost[0] + small_right_range,
                   rightmost[0]):
        tempright_img = img_jsw[:, x]
        testlen = len(tempright_img[tempright_img >= 200])
        right_X_list.append(x)
        right_length_list.append(testlen)

    right_avg = np.average(right_length_list) / (rightmost[0] - leftmost[0])
    right_std = np.std(right_length_list) / (rightmost[0] - leftmost[0])
    right_min = np.min(right_length_list) / (rightmost[0] - leftmost[0])

    gs_left_near_X = abs(gs_left_near_X)
    gs_right_near_X = abs(gs_right_near_X)
    gs_small_left_range = int((2 / 10) * (gs_right_near_X - gs_left_near_X))

    for x in range(gs_left_near_X,
                   gs_left_near_X + gs_small_left_range):
        temp_mask_left_img = mask_img[:, x]
        testlen = len(temp_mask_left_img[temp_mask_left_img >= 200])
        gs_left_X_list.append(x)
        gs_left_length_list.append(testlen)
    if len(left_length_list) == 0:
        gs_left_min = 0
        gs_left_avg = 0
        gs_left_std = 0
        gs_left_area = 0
    else:
        gs_left_avg = np.average(gs_left_length_list) / (rightmost[0] - leftmost[0])
        gs_left_min = np.min(gs_left_length_list) / (rightmost[0] - leftmost[0])
        gs_left_std = np.std(gs_left_length_list) / (rightmost[0] - leftmost[0])
        gs_left_area = np.sum(gs_left_length_list) / (rightmost[0] - leftmost[0])

    gs_large_right_range = int((8 / 10) * (gs_right_near_X - gs_left_near_X))
    gs_right_min_X = gs_left_near_X + gs_large_right_range

    for x in range(gs_left_near_X + gs_large_right_range,
                   gs_right_near_X):
        temp_mask_right_img = mask_img[:, x]
        testlen = len(temp_mask_right_img[temp_mask_right_img >= 200])
        gs_right_X_list.append(x)
        gs_right_length_list.append(testlen)
    # print(right_length_list)
    if len(right_length_list) == 0:
        gs_right_min = 0
        gs_right_avg = 0
        gs_right_std = 0
        gs_right_area = 0
    else:
        gs_right_avg = np.average(gs_right_length_list) / (rightmost[0] - leftmost[0])
        gs_right_std = np.std(gs_right_length_list) / (rightmost[0] - leftmost[0])
        gs_right_min = np.min(gs_right_length_list) / (rightmost[0] - leftmost[0])
        gs_right_area = np.sum(gs_right_length_list) / (rightmost[0] - leftmost[0])

    # cv2.line(mask_img, (gs_left_near_X, 0), (gs_left_near_X, mask_img.shape[1]),
    #          (255, 0, 0), 5)
    # cv2.line(mask_img, (gs_left_near_X + gs_small_left_range, 0), (gs_left_near_X + gs_small_left_range, mask_img.shape[1]),
    #          (255, 0, 0), 5)
    # cv2.line(mask_img, (gs_right_near_X, 0), (gs_right_near_X, mask_img.shape[1]),
    #          (255, 0, 0), 5)
    # cv2.line(mask_img, (gs_left_near_X + gs_large_right_range, 0), (gs_left_near_X + gs_large_right_range, mask_img.shape[1]),
    #          (255, 0, 0), 5)
    # cv2.line(mask_img, gs_fl, gs_fl, (0, 0, 255), 5)
    # cv2.line(mask_img, gs_fm, gs_fm, (0, 0, 255), 5)
    # cv2.line(mask_img, gs_tl, gs_tl, (0, 0, 255), 5)
    # cv2.line(mask_img, gs_tm, gs_tm, (0, 0, 255), 5)
    # # cv2.imwrite('./pilot_test/ ' + name, test_img)
    # cv2.imwrite('/app/threshold_test/' + 'mask_' + name, mask_img)
    # return left_avg, left_std, right_avg, right_std
    print("left avg : ", left_avg, 'left_std : ', left_std, 'right_avg : ', right_avg, 'right_std : ', right_std)
    print("gs left avg : ", gs_left_avg, 'gs left_std : ', gs_left_std, 'gs right_avg : ', gs_right_avg, 'gs right_std : ', gs_right_std)
    # print(gs_left_length_list)
    # print(gs_right_length_list)
    # cv2.line(test_img, (leftmost[0] + large_left_range, 0), (leftmost[0] + large_left_range, img_jsw.shape[1]), (255, 0, 0), 5)
    # cv2.line(test_img, (rightmost[0] + large_right_range, 0), (rightmost[0] + large_right_range, img_jsw.shape[1]), (255, 0, 0), 5)
    # cv2.imwrite('./pilot_test/ ' + name, test_img)
    # return left_avg, left_std, right_avg, right_std, left_min, right_min, -1, -1, -1, -1, -1, -1, -1, -1, -1, -1
    return left_avg, left_std, right_avg, right_std, left_min, right_min, gs_left_avg, gs_left_std, gs_right_avg, \
           gs_right_std, gs_left_min, gs_right_min, left_area, gs_left_area, right_area, gs_right_area
def abnormal_jsw_test(left_avg, right_avg, left_std, right_std, img, name, test_img):
    standard_left_X_rate = left_avg
    standard_right_X_rate = right_avg
    std_multiple = 4
    print(standard_left_X_rate)
    print(standard_right_X_rate)
    # print(img.shape)
    ret, img_jsw = cv2.threshold(img, int(0.2 * 255.), 255, 0)   # 이거 바꿔야함, keypoint를 하나라도 못찾는 경우의 threshold 0.1~0.5로 테스트하기. 이거랑 똑같이 바꿀 거 밑에 있음
    # for joint length
    ret, img_length = cv2.threshold(img, int(0.2 * 255.), 255, 0)
    # print(img_length.shape)
    img_length = cv2.cvtColor(img_length, cv2.COLOR_BGR2GRAY)

    _, contours, _ = cv2.findContours(img_length, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
    # left XY, right XY in joint object
    leftmost = contours[0][contours[0][:, :, 0].argmin()][0]
    rightmost = contours[0][contours[0][:, :, 0].argmax()][0]
    for c in contours:
        temp_left = c[c[:, :, 0].argmin()][0]
        temp_right = c[c[:, :, 0].argmax()][0]
        if leftmost[0] > temp_left[0]: leftmost = temp_left
        if rightmost[0] < temp_right[0]: rightmost = temp_right

    templrx = int(((rightmost[0] - leftmost[0]) / 2) + leftmost[0])
    # img_jsw = cv2.cvtColor(img_jsw, cv2.COLOR_GRAY2BGR)
    # a1 = int((templrx - leftmost[0]) * 0.1 + leftmost[0])
    # a4 = int((templrx - leftmost[0]) * 0.6 + leftmost[0])
    # b1 = int((rightmost[0] - templrx) * 0.4 + templrx)
    # b4 = int((rightmost[0] - templrx) * 0.9 + templrx)
    # if abs(leftmost[0] - int(img_jsw.shape[1] * standard_left_X_rate)) > img_jsw.shape[1] * (
    #         std_multiple * left_std):
    #     left_X = int(img_jsw.shape[1] * standard_left_X_rate)
    #     # left_X = leftmost[0]
    # else:
    #     left_X = leftmost[0]
    # if abs(rightmost[0] - int(img_jsw.shape[1] * standard_right_X_rate)) > img_jsw.shape[1] * (
    #         std_multiple * right_std):
    #     right_X = int(img_jsw.shape[1] * standard_right_X_rate)
    #     # right_X = rightmost[0]
    # else:
    #     right_X = rightmost[0]
    # left extraction & probe left joint space min, joint space min's average
    left_width_min = img_jsw.shape[0]
    left_X_list = []
    left_length_list = []
    # for x in range(left_X, int(img_jsw.shape[1] * 0.4)):
    # small_left_range = int((2 / 5) * (img_jsw.shape[1] * 0.4 - left_X))
    # cv2.line(img_jsw, (leftmost[0], 0), (leftmost[0], img_jsw.shape[1]), (255, 0, 0), 5)
    # cv2.line(img_jsw, (rightmost[0], 0), (rightmost[0], img_jsw.shape[1]), (255, 0, 0), 5)
    # cv2.imwrite('./oai_line_test/ ' + name, img_jsw)
    large_left_range = int((4 / 10) * ((rightmost[0] - leftmost[0]) * 0.4))
    for x in range(leftmost[0],
                   leftmost[0] + large_left_range):
        templeft_img = img_jsw[:, x]
        testlen = len(templeft_img[templeft_img >= 200])
        left_X_list.append(x)
        left_length_list.append(testlen)

    # left_avg = np.average(left_length_list) / (right_X - left_X)
    # left_min = np.min(left_length_list) / (right_X - left_X)
    # left_std = np.std(left_length_list) / (right_X - left_X)
    left_avg = np.average(left_length_list) / (rightmost[0] - leftmost[0])
    left_min = np.min(left_length_list) / (rightmost[0] - leftmost[0])
    left_std = np.std(left_length_list) / (rightmost[0] - leftmost[0])
    # print(left_avg)
    # # print(left_std)
    # left_temp_length_list = []
    # left_temp_X_list = []
    # for idx, left_x_length in enumerate(left_min_list):
    #     if abs(left_avg - left_x_length) <= left_std:
    #         left_temp_length_list.append(left_x_length)
    #         left_temp_X_list.append(left_min_X_list[idx])
    # # print(left_temp_length_list)
    # print('left min length :', np.average(left_temp_length_list) / abs(right_X - left_X))
    right_width_min = img_jsw.shape[0]
    right_X_list = []
    right_length_list = []
    # print(right_X)
    # print(int((4 / 5) * (right_X - img_jsw.shape[1] * 0.6)))
    # for x in range(int(img_jsw.shape[1] * 0.6), right_X):
    small_right_range = int((6 / 10) * ((rightmost[0] - leftmost[0]) * 0.4))
    large_right_range = int((4 / 5) * (rightmost[0] - img_jsw.shape[1] * 0.6))

    for x in range(leftmost[0] + small_right_range,
                   rightmost[0]):
        tempright_img = img_jsw[:, x]
        testlen = len(tempright_img[tempright_img >= 200])
        right_X_list.append(x)
        right_length_list.append(testlen)

    right_avg = np.average(right_length_list) / (rightmost[0] - leftmost[0])
    right_std = np.std(right_length_list) / (rightmost[0] - leftmost[0])
    right_min = np.min(right_length_list) / (rightmost[0] - leftmost[0])
    # return left_avg, left_std, right_avg, right_std
    # cv2.line(test_img, (leftmost[0] + large_left_range, 0), (leftmost[0] + large_left_range, img_jsw.shape[1]), (255, 0, 0), 5)
    # cv2.line(test_img, (rightmost[0] + large_right_range, 0), (rightmost[0] + large_right_range, img_jsw.shape[1]), (255, 0, 0), 5)
    # cv2.imwrite('./pilot_test/ ' + name, test_img)
    return left_avg, left_std, right_avg, right_std, left_min, right_min
def calculator_jsw_test_Threshold(left_avg, right_avg, left_std, right_std, img, keypoint_list, name, test_img, Threshold):
    abnormal_flag = False
    for op_area, op_xy in keypoint_list:
        # print(op_area)
        # print(op_xy)
        if op_area == 0:
            abnormal_flag = True
        if op_area == '0':
            fl = op_xy
        elif op_area == '1':
            fm = op_xy
        elif op_area == '2':
            tl = op_xy
        elif op_area == '3':
            tm = op_xy
    if abnormal_flag:
        return abnormal_jsw_test_Threshold(left_avg, right_avg, left_std, right_std, img, name, test_img, Threshold)
    standard_left_X_rate = left_avg
    standard_right_X_rate = right_avg
    std_multiple = 4
    # cv2.imwrite('/app/img_test/' + name, img)
    # print(img.shape)
    print(int(Threshold * 255.))
    ret, img_jsw = cv2.threshold(img, int(Threshold * 255.), 255, 0)
    # for joint length
    ret, img_length = cv2.threshold(img, int(Threshold * 255.), 255, 0)


    img_length = cv2.cvtColor(img_length, cv2.COLOR_BGR2GRAY)
    print(img_length.shape)
    if fl[0] < tl[0]:
        left_near_X = tl[0]
    else:
        left_near_X = fl[0]
    if fm[0] > tm[0]:
        right_near_X = tm[0]
    else:
        right_near_X = fm[0]

    left_near_X = abs(left_near_X)
    right_near_X = abs(right_near_X)
    print(left_near_X)
    print(right_near_X)

    left_X_list = []
    left_length_list = []

    small_left_range = int((2 / 10) * (right_near_X - left_near_X))

    for x in range(left_near_X,
                   left_near_X + small_left_range):
        templeft_img = img_jsw[:, x]
        testlen = len(templeft_img[templeft_img >= 200])
        left_X_list.append(x)
        left_length_list.append(testlen)
    if len(left_length_list) == 0:
        left_min = 0
        left_avg = 0
        left_std = 0
    else:
        left_avg = np.average(left_length_list) / (right_near_X - left_near_X)
        left_min = np.min(left_length_list) / (right_near_X - left_near_X)
        left_std = np.std(left_length_list) / (right_near_X - left_near_X)


    right_X_list = []
    right_length_list = []

    large_right_range = int((8 / 10) * (right_near_X - left_near_X))

    for x in range(left_near_X + large_right_range,
                   right_near_X):
        tempright_img = img_jsw[:, x]
        testlen = len(tempright_img[tempright_img >= 200])
        right_X_list.append(x)
        right_length_list.append(testlen)
    # print(right_length_list)
    if len(right_length_list) == 0:
        right_min = 0
        right_avg = 0
        right_std = 0
    else:
        right_avg = np.average(right_length_list) / (right_near_X - left_near_X)
        right_std = np.std(right_length_list) / (right_near_X - left_near_X)
        right_min = np.min(right_length_list) / (right_near_X - left_near_X)

    compare_list = ['021_AP0000L_joint.jpg', '091_FLX0000L_joint.jpg', '070_AP_20190000R_joint.jpg', '049_AP0000L_joint.jpg',
                    '062_EXAP_20190000R_joint.jpg', '057_AP_20190000L_joint.jpg', '021_AP0000R_joint.jpg', '053_EXAP0000L_joint.jpg',
                    '023_AP0000R_joint.jpg', '052_FLX0000L_joint.jpg']
    # if name in compare_list:
    cv2.line(img_jsw, (left_near_X + small_left_range, 0), (left_near_X + small_left_range, img_jsw.shape[1]),
             (255, 0, 0), 5)
    cv2.line(img_jsw, (left_near_X + large_right_range, 0), (left_near_X + large_right_range, img_jsw.shape[1]), (255, 0, 0), 5)
    cv2.line(img_jsw, (left_near_X, 0), (left_near_X, img_jsw.shape[1]),
             (255, 0, 0), 5)
    cv2.line(img_jsw, (right_near_X, 0), (right_near_X, img_jsw.shape[1]),
             (255, 0, 0), 5)
    cv2.line(img_jsw, fl, fl, (0, 0, 255), 5)
    cv2.line(img_jsw, fm, fm, (0, 0, 255), 5)
    cv2.line(img_jsw, tl, tl, (0, 0, 255), 5)
    cv2.line(img_jsw, tm, tm, (0, 0, 255), 5)
    # cv2.imwrite('./pilot_test/ ' + name, test_img)
    cv2.imwrite('./theshold_test/' + one_file_name.replace('_joint.jpg', '') + '/' + str(T) + '_' + name, img_jsw)
    print("left avg : ", left_avg, 'left_std : ', left_std, 'right_avg : ', right_avg, 'right_std : ', right_std)
    return left_avg, left_std, right_avg, right_std, left_min, right_min
def json_to_str(json_name):
    #with open(os.path.join('/app2/pilot/key_img', json_name)) as f:
    with open(os.path.join('/app/pilot/key_img', json_name)) as f:
        json_ob = json.load(f)
    points = []
    for cnt in range(0, 4):
        # print(json_ob['shapes'][cnt])
        points.append((json_ob['shapes'][cnt]['label'], json_ob['shapes'][cnt]['points'][0][0], json_ob['shapes'][cnt]['points'][0][1]))
    return points
def calculator_jsw_test_Threshold_compare_gs(left_avg, right_avg, left_std, right_std, img, keypoint_list, name, test_img, Threshold):
    abnormal_flag = False
    for op_area, op_xy in keypoint_list:
        # print(op_area)
        # print(op_xy)
        if op_area == 0:
            abnormal_flag = True
        if op_area == '0':
            fl = op_xy
        elif op_area == '1':
            fm = op_xy
        elif op_area == '2':
            tl = op_xy
        elif op_area == '3':
            tm = op_xy
    if abnormal_flag:
        return abnormal_jsw_test_Threshold(left_avg, right_avg, left_std, right_std, img, name, test_img, Threshold)
    standard_left_X_rate = left_avg
    standard_right_X_rate = right_avg
    std_multiple = 4
    # cv2.imwrite('/app/img_test/' + name, img)
    # print(img.shape)
    points = json_to_str(name.replace('.jpg', '.json'))
    for point_info in points:
        if point_info[0] == 'fl':
            gs_fl = (int(point_info[1]), int(point_info[2]))
        elif point_info[0] == 'fm':
            gs_fm = (int(point_info[1]), int(point_info[2]))
        elif point_info[0] == 'tl':
            gs_tl = (int(point_info[1]), int(point_info[2]))
        elif point_info[0] == 'tm':
            gs_tm = (int(point_info[1]), int(point_info[2]))

#    mask_img = cv2.imread(os.path.join('/app2/pilot/img_mask', name.replace('.jpg', '.png')))
    mask_img = cv2.imread(os.path.join('/app/pilot/img_mask', name.replace('.jpg', '.png')))

    mask_img *= 255

    # print(int(Threshold * 255.))
    ret, img_jsw = cv2.threshold(img, int(Threshold * 255.), 255, 0)
    # for joint length
    ret, img_length = cv2.threshold(img, int(Threshold * 255.), 255, 0)
    # x_minus = (img_jsw.shape[1] - img_jsw.shape[0]) / 2
    # x_minus = int(x_minus)
    # img_jsw1 = img_jsw.copy()
    # if img_jsw.shape[1] == img_jsw.shape[0] + (x_minus * 2):
    #     img_jsw = cv2.copyMakeBorder(img_jsw, x_minus, x_minus, 0, 0, cv2.BORDER_CONSTANT)
    # else:
    #     img_jsw = cv2.copyMakeBorder(img_jsw, x_minus, x_minus + 1, 0, 0, cv2.BORDER_CONSTANT)
    #
    # img_jsw = cv2.resize(img_jsw, (mask_img.shape[1], mask_img.shape[0]))
    img_length = cv2.cvtColor(img_length, cv2.COLOR_BGR2GRAY)
    # print(img_length.shape)
    if fl[0] < tl[0]:
        left_near_X = tl[0]
    else:
        left_near_X = fl[0]
    if fm[0] > tm[0]:
        right_near_X = tm[0]
    else:
        right_near_X = fm[0]

    if gs_fl[0] < gs_tl[0]:
        gs_left_near_X = gs_tl[0]
    else:
        gs_left_near_X = gs_fl[0]
    if gs_fm[0] > gs_tm[0]:
        gs_right_near_X = gs_tm[0]
    else:
        gs_right_near_X = gs_fm[0]

    left_near_X = abs(left_near_X)
    right_near_X = abs(right_near_X)

    # left_near_X = int((left_near_X / img_jsw1.shape[1]) * img_jsw.shape[1])
    # right_near_X = int((right_near_X / img_jsw1.shape[1]) * img_jsw.shape[1])
    left_X_list = []
    left_length_list = []
    gs_left_X_list = []
    gs_left_length_list = []
    small_left_range = int((2 / 10) * (right_near_X - left_near_X))
    small_left_side_range = int((1 / 20) * (right_near_X - left_near_X))
    for x in range(left_near_X + small_left_side_range,
                   left_near_X + small_left_range):
        templeft_img = img_jsw[:, x]
        testlen = len(templeft_img[templeft_img >= 200])
        left_X_list.append(x)
        left_length_list.append(testlen)
    if len(left_length_list) == 0:
        left_min = 0
        left_avg = 0
        left_std = 0
        left_area = 0
    else:
        left_avg = np.average(left_length_list) / (right_near_X - left_near_X)
        left_min = np.min(left_length_list) / (right_near_X - left_near_X)
        left_std = np.std(left_length_list) / (right_near_X - left_near_X)
        left_area = np.sum(left_length_list) / (right_near_X - left_near_X)

    gs_left_near_X = abs(gs_left_near_X)
    gs_right_near_X = abs(gs_right_near_X)
    gs_small_left_range = int((2 / 10) * (gs_right_near_X - gs_left_near_X))
    gs_small_left_side_range = int((1 / 20) * (gs_right_near_X - gs_left_near_X))
    for x in range(gs_left_near_X + gs_small_left_side_range,
                   gs_left_near_X + gs_small_left_range):
        temp_mask_left_img = mask_img[:, x]
        testlen = len(temp_mask_left_img[temp_mask_left_img >= 200])
        gs_left_X_list.append(x)
        gs_left_length_list.append(testlen)
    if len(left_length_list) == 0:
        gs_left_min = 0
        gs_left_avg = 0
        gs_left_std = 0
        gs_left_area = 0
    else:
        gs_left_avg = np.average(gs_left_length_list) / (right_near_X - left_near_X)
        gs_left_min = np.min(gs_left_length_list) / (right_near_X - left_near_X)
        gs_left_std = np.std(gs_left_length_list) / (right_near_X - left_near_X)
        gs_left_area = np.sum(gs_left_length_list) / (right_near_X - left_near_X)


    right_X_list = []
    right_length_list = []

    gs_right_X_list = []
    gs_right_length_list = []
    large_right_range = int((8 / 10) * (right_near_X - left_near_X))
    large_right_side_range = int((1 / 20) * (right_near_X - left_near_X))
    for x in range(left_near_X + large_right_range,
                   right_near_X - large_right_side_range):
        tempright_img = img_jsw[:, x]
        testlen = len(tempright_img[tempright_img >= 200])
        right_X_list.append(x)
        right_length_list.append(testlen)
    # print(right_length_list)
    if len(right_length_list) == 0:
        right_min = 0
        right_avg = 0
        right_std = 0
        right_area = 0
    else:
        right_avg = np.average(right_length_list) / (right_near_X - left_near_X)
        right_std = np.std(right_length_list) / (right_near_X - left_near_X)
        right_min = np.min(right_length_list) / (right_near_X - left_near_X)
        right_area = np.sum(right_length_list) / (right_near_X - left_near_X)
    gs_large_right_range = int((8 / 10) * (gs_right_near_X - gs_left_near_X))
    gs_right_min_X = gs_left_near_X + gs_large_right_range

    for x in range(gs_left_near_X + gs_large_right_range,
                   gs_right_near_X):
        temp_mask_right_img = mask_img[:, x]
        testlen = len(temp_mask_right_img[temp_mask_right_img >= 200])
        gs_right_X_list.append(x)
        gs_right_length_list.append(testlen)
    # print(right_length_list)
    if len(right_length_list) == 0:
        gs_right_min = 0
        gs_right_avg = 0
        gs_right_std = 0
        gs_right_area = 0
    else:
        gs_right_avg = np.average(gs_right_length_list) / (right_near_X - left_near_X)
        gs_right_std = np.std(gs_right_length_list) / (right_near_X - left_near_X)
        gs_right_min = np.min(gs_right_length_list) / (right_near_X - left_near_X)
        gs_right_area = np.sum(gs_right_length_list) / (right_near_X - left_near_X)

    cv2.line(img_jsw, (left_near_X + small_left_range, 0), (left_near_X + small_left_range, img_jsw.shape[1]),
             (255, 0, 0), 5)
    cv2.line(img_jsw, (left_near_X + large_right_range, 0), (left_near_X + large_right_range, img_jsw.shape[1]), (255, 0, 0), 5)
    cv2.line(img_jsw, (left_near_X + small_left_side_range, 0), (left_near_X + small_left_side_range, img_jsw.shape[1]),
             (255, 0, 0), 5)
    cv2.line(img_jsw, (right_near_X - large_right_side_range, 0), (right_near_X - large_right_side_range, img_jsw.shape[1]),
             (255, 0, 0), 5)
    cv2.line(img_jsw, fl, fl, (0, 0, 255), 5)
    cv2.line(img_jsw, fm, fm, (0, 0, 255), 5)
    cv2.line(img_jsw, tl, tl, (0, 0, 255), 5)
    cv2.line(img_jsw, tm, tm, (0, 0, 255), 5)
    # cv2.imwrite('./pilot_test/ ' + name, test_img)
    # print('./theshold_test/' + '_' + name)
    #cv2.imwrite('/app/threshold_test/' + name, img_jsw)
    # cv2.line(mask_img, (gs_left_near_X, 0), (gs_left_near_X, mask_img.shape[1]),
    #          (255, 0, 0), 5)
    # cv2.line(mask_img, (gs_left_near_X + gs_small_left_range, 0), (gs_left_near_X + gs_small_left_range, mask_img.shape[1]),
    #          (255, 0, 0), 5)
    # cv2.line(mask_img, (gs_right_near_X, 0), (gs_right_near_X, mask_img.shape[1]),
    #          (255, 0, 0), 5)
    # cv2.line(mask_img, (gs_left_near_X + gs_large_right_range, 0), (gs_left_near_X + gs_large_right_range, mask_img.shape[1]),
    #          (255, 0, 0), 5)
    # cv2.line(mask_img, gs_fl, gs_fl, (0, 0, 255), 5)
    # cv2.line(mask_img, gs_fm, gs_fm, (0, 0, 255), 5)
    # cv2.line(mask_img, gs_tl, gs_tl, (0, 0, 255), 5)
    # cv2.line(mask_img, gs_tm, gs_tm, (0, 0, 255), 5)
    # cv2.imwrite('./pilot_test/ ' + name, test_img)
    cv2.imwrite('/app/mask_test/' + 'mask_' + name, img_jsw)

    # print("left avg : ", left_avg, 'left_std : ', left_std, 'right_avg : ', right_avg, 'right_std : ', right_std)
    return left_avg, left_std, right_avg, right_std, left_min, right_min, gs_left_avg, gs_left_std, gs_right_avg, \
           gs_right_std, gs_left_min, gs_right_min, left_area, gs_left_area, right_area, gs_right_area
# analysis (left average, standard deviation)
def calculator_jsw_test(left_avg, right_avg, left_std, right_std, img, keypoint_list, name, test_img):
    abnormal_flag = False
    for op_area, op_xy in keypoint_list:
        # print(op_area)
        # print(op_xy)
        if op_area == 0:
            abnormal_flag = True
        if op_area == '0':
            fl = op_xy
        elif op_area == '1':
            fm = op_xy
        elif op_area == '2':
            tl = op_xy
        elif op_area == '3':
            tm = op_xy
    if abnormal_flag:
        return abnormal_jsw_test(left_avg, right_avg, left_std, right_std, img, name, test_img)
    standard_left_X_rate = left_avg
    standard_right_X_rate = right_avg
    std_multiple = 4

    # print(img.shape)
    ret, img_jsw = cv2.threshold(img, int(0.2 * 255.), 255, 0)
    # for joint length
    ret, img_length = cv2.threshold(img, int(0.2 * 255.), 255, 0)

    # cv2.imwrite('/app/oai_line_test/' + name, img_length)
    img_length = cv2.cvtColor(img_length, cv2.COLOR_BGR2GRAY)
    print(img_length.shape)
    if fl[0] < tl[0]:
        left_near_X = tl[0]
    else:
        left_near_X = fl[0]
    if fm[0] > tm[0]:
        right_near_X = tm[0]
    else:
        right_near_X = fm[0]

    left_near_X = abs(left_near_X)
    right_near_X = abs(right_near_X)
    print(left_near_X)
    print(right_near_X)

    left_X_list = []
    left_length_list = []

    small_left_range = int((2 / 10) * (right_near_X - left_near_X))

    for x in range(left_near_X,
                   left_near_X + small_left_range):
        templeft_img = img_jsw[:, x]
        testlen = len(templeft_img[templeft_img >= 200])
        left_X_list.append(x)
        left_length_list.append(testlen)
    if len(left_length_list) == 0:
        left_min = 0
        left_avg = 0
        left_std = 0
    else:
        left_avg = np.average(left_length_list) / (right_near_X - left_near_X)
        left_min = np.min(left_length_list) / (right_near_X - left_near_X)
        left_std = np.std(left_length_list) / (right_near_X - left_near_X)
    cv2.line(test_img, (left_near_X + small_left_range, 0), (left_near_X + small_left_range, img_jsw.shape[1]), (255, 0, 0), 5)

    right_X_list = []
    right_length_list = []

    large_right_range = int((8 / 10) * (right_near_X - left_near_X))

    for x in range(left_near_X + large_right_range,
                   right_near_X):
        tempright_img = img_jsw[:, x]
        testlen = len(tempright_img[tempright_img >= 200])
        right_X_list.append(x)
        right_length_list.append(testlen)
    # print(right_length_list)
    if len(right_length_list) == 0:
        right_min = 0
        right_avg = 0
        right_std = 0
    else:
        right_avg = np.average(right_length_list) / (right_near_X - left_near_X)
        right_std = np.std(right_length_list) / (right_near_X - left_near_X)
        right_min = np.min(right_length_list) / (right_near_X - left_near_X)
    # cv2.line(test_img, (left_near_X + large_right_range, 0), (left_near_X + large_right_range, img_jsw.shape[1]), (255, 0, 0), 5)
    # cv2.line(test_img, (left_near_X, 0), (left_near_X, img_jsw.shape[1]),
    #          (255, 0, 0), 5)
    # cv2.line(test_img, (right_near_X, 0), (right_near_X, img_jsw.shape[1]),
    #          (255, 0, 0), 5)
    # cv2.line(test_img, fl, fl, (0, 0, 255), 5)
    # cv2.line(test_img, fm, fm, (0, 0, 255), 5)
    # cv2.line(test_img, tl, tl, (0, 0, 255), 5)
    # cv2.line(test_img, tm, tm, (0, 0, 255), 5)
    # cv2.imwrite('./pilot_test/ ' + name, test_img)

    return left_avg, left_std, right_avg, right_std, left_min, right_min


def load_excel_gs_pilot():
#    workbook = load_workbook('./pilot.xlsx')
    workbook = load_workbook('/app/pilot.xlsx')

    names = []
    kls = []
    # fls = []
    # fms = []
    # tls = []
    # tms = []
    # ws = ['Sheet1']
    ws = ['Sheet1']
    for sh_num in ws:
        sheet = workbook[sh_num]
        for c1, c2 in sheet[sheet.dimensions]:
            print(c1.value)
            names.append(c1.value)
            kls.append(c2.value)
            # fls.append(c3.value)
            # fms.append(c4.value)
            # tls.append(c5.value)
            # tms.append(c6.value)
    # return names, kls, fls, fms, tls, tms
    return names, kls


def load_excel_gs():
    workbook = load_workbook('./meta_data_bilateral_v3.xlsx')
    names = []
    kls = []
    ws = ['Sheet1']
    for sh_num in ws:
        sheet = workbook[sh_num]
        for c1, c2, c3, c4, c5, c6, c7, c8, c9, c10, c11, c12, c13, c14, c15, c16, c17, c18, c19, c20, c21, c22, c23, c24, c25, c26 in sheet[sheet.dimensions]:
            if c2.value == '2':
                names.append(c5.value + 'L' + c3.value + '.jpg')
            else:
                names.append(c5.value + 'R' + c3.value + '.jpg')
            kls.append(c8.value)
    names.pop(0)
    kls.pop(0)
    return names, kls


def load_excel_point(order):
    if order == 'train':
        workbook = load_workbook('0616_total_op_point.xlsx')
    elif order == 'test':
        workbook = load_workbook('0702_validation_op_point.xlsx')
    else:
        workbook = load_workbook('/app/0709_pilot_op_point.xlsx')

    names = []
    keypoints = []
    fls = []
    fms = []
    tls = []
    tms = []

    # ws = ['0', '1', '2', '3', '4']
    ws = ['Sheet1']
    for sh_num in ws:
        sheet = workbook[sh_num]
        for c1, c2, c3, c4, c5 in sheet[sheet.dimensions]:
            # lists.append((c1.value, c2.value))

            temp_name = c1.value
            # print(temp_name)
            # if c1.value[7] == 'L':
            #     temp_name = temp_name[0:7] + 'L' + c1.value[7:9] + c1.value[10:]
            # else:
            #     temp_name = temp_name[0:7] + 'R' + c1.value[7:9] + c1.value[10:]

            print(temp_name)
            names.append(c1.value)
            fls.append(c2.value)
            fms.append(c3.value)
            tls.append(c4.value)
            tms.append(c5.value)
    # names.pop(0)
    # # keypoints.pop(0)
    # fls.pop(0)
    # fms.pop(0)
    # tls.pop(0)
    # tms.pop(0)
    return names, fls, fms, tls, tms
# keypoint list -> hrnet predct
def calculator_jsw(left_avg, right_avg, left_std, right_std, img, keypoint_list):
    if len(keypoint_list) == 4:
        tl_xmin, tl_ymin = keypoint_list[2]
        tm_xmin, tm_ymin = keypoint_list[3]

    standard_left_X_rate = left_avg
    standard_right_X_rate = right_avg
    std_multiple = 4
    save_length_list = []
    temp_length = []
    # print(img.shape)
    ret, img_jsw = cv2.threshold(img, int(0.3 * 255.), 255, 0)
    # for joint length
    ret, img_length = cv2.threshold(img, int(0.3 * 255.), 255, 0)
    # print(img_length.shape)
    # img_length = cv2.cvtColor(img_length, cv2.COLOR_BGR2GRAY)
    contours, _ = cv2.findContours(img_length, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
    # left XY, right XY in joint object
    leftmost = contours[0][contours[0][:, :, 0].argmin()][0]
    rightmost = contours[0][contours[0][:, :, 0].argmax()][0]
    for c in contours:
        temp_left = c[c[:, :, 0].argmin()][0]
        temp_right = c[c[:, :, 0].argmax()][0]
        if leftmost[0] > temp_left[0]: leftmost = temp_left
        if rightmost[0] < temp_right[0]: rightmost = temp_right

    templrx = int(((rightmost[0] - leftmost[0]) / 2) + leftmost[0])
    img_jsw = cv2.cvtColor(img_jsw, cv2.COLOR_GRAY2BGR)
    # a1 = int((templrx - leftmost[0]) * 0.1 + leftmost[0])
    # a4 = int((templrx - leftmost[0]) * 0.6 + leftmost[0])
    # b1 = int((rightmost[0] - templrx) * 0.4 + templrx)
    # b4 = int((rightmost[0] - templrx) * 0.9 + templrx)
    if abs(leftmost[0] - int(img_jsw.shape[1] * standard_left_X_rate)) > img_jsw.shape[1] * (
            std_multiple * left_std):
        left_X = int(img_jsw.shape[1] * standard_left_X_rate)
        # left_X = leftmost[0]
    else:
        left_X = leftmost[0]
    if abs(rightmost[0] - int(img_jsw.shape[1] * standard_right_X_rate)) > img_jsw.shape[1] * (
            std_multiple * right_std):
        right_X = int(img_jsw.shape[1] * standard_right_X_rate)
        # right_X = rightmost[0]
    else:
        right_X = rightmost[0]
    # left extraction & probe left joint space min, joint space min's average
    left_width_min = img_jsw.shape[0]
    left_X_list = []
    left_length_list = []
    # for x in range(left_X, int(img_jsw.shape[1] * 0.4)):
    # small_left_range = int((2 / 5) * (img_jsw.shape[1] * 0.4 - left_X))
    # 4 개가 아닌 경우
    if not len(keypoint_list) == 4:
        tm_xmin = right_X
        tl_xmin = left_X
    large_left_range = int((4 / 10) * ((tm_xmin - tl_xmin) * 0.4))
    for x in range(left_X,
                   left_X + large_left_range):
        templeft_img = img_jsw[:, x]
        testlen = len(templeft_img[templeft_img >= 200])
        left_X_list.append(x)
        left_length_list.append(testlen)

    # left_avg = np.average(left_length_list) / (right_X - left_X)
    # left_min = np.min(left_length_list) / (right_X - left_X)
    # left_std = np.std(left_length_list) / (right_X - left_X)
    left_avg = np.average(left_length_list) / (tm_xmin - tl_xmin)
    left_min = np.min(left_length_list) / (tm_xmin - tl_xmin)
    left_std = np.std(left_length_list) / (tm_xmin - tl_xmin)
    # print(left_avg)
    # # print(left_std)
    # left_temp_length_list = []
    # left_temp_X_list = []
    # for idx, left_x_length in enumerate(left_min_list):
    #     if abs(left_avg - left_x_length) <= left_std:
    #         left_temp_length_list.append(left_x_length)
    #         left_temp_X_list.append(left_min_X_list[idx])
    # # print(left_temp_length_list)
    # print('left min length :', np.average(left_temp_length_list) / abs(right_X - left_X))
    right_width_min = img_jsw.shape[0]
    right_X_list = []
    right_length_list = []
    # print(right_X)
    # print(int((4 / 5) * (right_X - img_jsw.shape[1] * 0.6)))
    # for x in range(int(img_jsw.shape[1] * 0.6), right_X):
    small_right_range = int((6 / 10) * ((tm_xmin - tl_xmin) * 0.4))
    large_right_range = int((4 / 5) * (right_X - img_jsw.shape[1] * 0.6))

    for x in range(left_X + small_right_range,
                   right_X):
        tempright_img = img_jsw[:, x]
        testlen = len(tempright_img[tempright_img >= 200])
        right_X_list.append(x)
        right_length_list.append(testlen)

    right_avg = np.average(right_length_list) / (tm_xmin - tl_xmin)
    right_std = np.std(right_length_list) / (tm_xmin - tl_xmin)
    right_min = np.min(right_length_list) / (tm_xmin - tl_xmin)
    # return left_avg, left_std, right_avg, right_std
    return left_avg, left_std, right_avg, right_std

# length_list, left_std, right_std, left_avg, right_avg = contour(os.listdir(filepath))
# contour_img(left_avg, right_avg, left_std, right_std, length_list)
# one_contour_img(left_avg, right_avg, left_std, right_std, length_list)
left_std = 0.017413596170162218
right_std = 0.018735075618898804
left_avg = 0.08525850101915945
right_avg = 0.9105576164087643
# OAI_contour_img_std(left_avg, right_avg, left_std, right_std)

# OAI_contour_img_std_test(left_avg, right_avg, left_std, right_std)
#OAI_contour_img_5_group(left_avg, right_avg, left_std, right_std)
#
# # TODO
# OAI_contour_img_std_test_bio_info(left_avg, right_avg, left_std, right_std)
order = 'pilot' # train, test, pilot
# Threshold = [0.3, 0.32, 0.34, 0.36, 0.38, 0.40, 0.42, 0.44, 0.46, 0.48, 0.5]
Threshold = [0.573, 0.574, 0.575]
#Threshold = [0.42]

names, p_fls, p_fms, p_tls, p_tms = load_excel_point(order)
if order == 'train':
    gs_names, kls = load_excel_gs()
    path = '/app2/0616_new_train5555_img_half_augment_add_img_rate_data_low_learning_rate/total_set_result/y_pred/positive'
    compare_path = '/app/keras_retinanet/keras_retinanet/bin/oai_total/joint'

elif order == 'test':
    gs_names, kls = load_excel_gs()
    path = '/app2/0629_new_img_half_augment_add_img_rate_data_low_learning_rate/total_validation_result/y_pred/positive'
    compare_path = '/app/keras_retinanet/keras_retinanet/bin/oai_validation/joint'
else:
    # gs_names, kls, fls, fms, tls, tms = load_excel_gs_pilot()
    gs_names, kls = load_excel_gs_pilot()
#    path = '/app2/0608_new_train5555_img_half_augment_add_img_rate_data/pilot_result/y_pred/positive'
    # path = '/app2/0604_new_train5550_img_half_augment_add_img_rate/oai_test_result/y_pred/positive'
    path = '/app/0802_new_img_half_augment_data_relabeled_redraw_data/pilot_result_new/y_pred/positive'
#    compare_path = '/app/keras_retinanet/keras_retinanet/bin/oai_pilot/joint'
    compare_path = '/app/pilot/img'

# print(gs_names)
# path = '/app/keras_retinanet/keras_retinanet/bin/oai_total/0512_add_data_original_learning_separate_addtrain_gamma_correction/img_result/y_pred/positive'
# path = '/app2/0608_new_train5555_img_half_augment_add_img_rate_data/total_set_result/y_pred/positive'


#save_path = './oai_total_matching_img'
save_path = '/app/oai_total_matching_img'

print("Threshold  : ", Threshold)

#************* 여기서부터 수정한 것 *************
#gs_val_name = os.listdir('/app2/pilot/key_img')
gs_val_name = os.listdir('/app/pilot/key_img')

comfirm_img_names = []
for gs_name in gs_val_name:
    if '.json' in gs_name:
        comfirm_img_names.append(gs_name.split('.json')[0])
# print(comfirm_img_names)
for T in Threshold:
    temp_list = []
    # print(compare_path)
    for one_file_name in os.listdir(compare_path):
        if 'oai_total' in compare_path:
            temp_name = one_file_name[0:10] + '.jpg'
            print(temp_name)
        elif 'oai_pilot' in compare_path:
            temp_name = one_file_name[0:10] + '.jpg'
            # print(temp_name)
        else:
#            temp_name = one_file_name[0:10] + '.jpg'
            temp_name = one_file_name[0:10] + '.jpg'
            #print(temp_name)


            # if one_file_name[9] == 'L':
            #     temp_name = one_file_name[0:7] + 'L' + one_file_name[7:9] + '.jpg' # one_file_name[7:10] + "_joint.jpg"
            #     # one_file_name = one_file_name[0:7] + one_file_name[8:9] + 'L' + '_joint.jpg'
            # else:
            #     temp_name = one_file_name[0:7] + 'R' + one_file_name[7:9] + '.jpg'  # one_file_name[7:10] + "_joint.jpg"
            #     # one_file_name = one_file_name[0:7] + one_file_name[8:9] + 'R' + '_joint.jpg'
            # print(origin_name)
            # print(temp_name)
        # if '900782748R_joint.jpg' in one_file_name:
        # if one_file_name.split('_joint.jpg')[0] not in comfirm_img_names:
        #     continue
        origin_img = cv2.imread(os.path.join(compare_path, one_file_name))
        addy_value = one_file_name
        addy_value = addy_value[:-4]+'.jpg'
        print(addy_value)
        print("origin_img : ",os.path.join(compare_path, one_file_name))
        print("analysis_img : ",os.path.join(path, addy_value))
        analysis_img = cv2.imread(os.path.join(path, addy_value))
        print(origin_img.shape)
        print(analysis_img.shape)

        x_minus = (analysis_img.shape[1] - analysis_img.shape[0]) / 2
        x_minus = int(x_minus)
        x_original_1 = analysis_img.copy()
        if analysis_img.shape[1] == analysis_img.shape[0] + (x_minus * 2):
            analysis_img = cv2.copyMakeBorder(analysis_img, x_minus, x_minus, 0, 0, cv2.BORDER_CONSTANT)
        else:
            analysis_img = cv2.copyMakeBorder(analysis_img, x_minus, x_minus + 1, 0, 0, cv2.BORDER_CONSTANT)

        analysis_img = cv2.resize(analysis_img, (origin_img.shape[1], origin_img.shape[0]))
        temp_img = origin_img.copy()
            # print(analysis_img.shape)
        # left_superimposed_img = cv2.addWeighted(analysis_img, 0.8, origin_img, 0.4, 0)
        # cv2.imwrite(os.path.join(save_path, 'test_overlay_' + one_file_name), left_superimposed_img)
        print('--------------------------------------------')
        # print(temp_name)
        # print(gs_names)
        if 'pilot' in order:
            temp_name = one_file_name
        print(1)
        print(one_file_name)
        if temp_name in gs_names:
            # print(gs_names.index(temp_name))
            print("???")
            kl = kls[gs_names.index(temp_name)]
        else:
            kl = -1
        tempN = one_file_name
        #if 'oai_pilot' in compare_path:
        if 'pilot' in compare_path:
            # one_file_name = one_file_name[0:8] + one_file_name[8:11] + '_joint.jpg'
            one_file_name = one_file_name[:-4] + '.jpg'
            print(one_file_name)
        else:
            one_file_name = one_file_name[0:8] + one_file_name[8:10] + '_joint.jpg'

        # print(one_file_name)
        # print(compare_img.shape)
        # print(names)
        # print(keypoints[names.index(temp_name)])
        # print(analysis_img.shape)
        keypoints = []
        if not p_fls[names.index(one_file_name)]:
            # print("?")
            keypoints.append((0, 0))
        else:
            # print(fls[names.index(one_file_name)])
            op_area = p_fls[names.index(one_file_name)].replace('(', '').replace(')', '').replace(' ', '').split(',')[0] # op area
            fl_x = int(p_fls[names.index(one_file_name)].replace('(', '').replace(')', '').replace(' ', '').split(',')[1]) # x coordinate
            fl_y = int(p_fls[names.index(one_file_name)].replace('(', '').replace(')', '').replace(' ', '').split(',')[2]) # y coordinate
            if analysis_img.shape[1] < fl_x or analysis_img.shape[0] < fl_y:
                continue
            # cv2.line(temp_img, (fl_x, fl_y), (fl_x, fl_y), (0, 0, 255), 5)
            keypoints.append((op_area, (fl_x, fl_y)))
            # keypoints.append((int(fls[names.index(one_file_name)].split(')')[0].split('(')[1].split(',')[0]),
            #                   int(fls[names.index(one_file_name)].split(')')[0].split('(')[1].split(',')[1].split(' ')[1]),
            #                   int(fls[names.index(one_file_name)].split(')')[0].split('(')[1].split(',')[1].split(' ')[2])))
            # print(fls[names.index(temp_name)].split(')')[0].split('(')[1].split(',')[0])
            # print(fls[names.index(temp_name)].split(')')[0].split('(')[1].split(',')[1].split(' ')[1])
            # print(type(fls[names.index(temp_name)]))
        if not p_fms[names.index(one_file_name)]:
            # print("?")
            keypoints.append((0, 0))
        else:
            op_area = p_fms[names.index(one_file_name)].replace('(', '').replace(')', '').replace(' ', '').split(',')[
                0]  # op area
            fm_x = int(p_fms[names.index(one_file_name)].replace('(', '').replace(')', '').replace(' ', '').split(',')[
                           1])  # x coordinate
            fm_y = int(p_fms[names.index(one_file_name)].replace('(', '').replace(')', '').replace(' ', '').split(',')[
                           2])  # y coordinate
            if analysis_img.shape[1] < fm_x or analysis_img.shape[0] < fm_y:
                continue
            # cv2.line(temp_img, (fm_x, fm_y), (fm_x, fm_y), (0, 0, 255), 5)
            keypoints.append((op_area, (fm_x, fm_y)))
            # print(type(fms[names.index(one_file_name)]))
            # keypoints.append((int(fms[names.index(one_file_name)].split(')')[0].split('(')[1].split(',')[0]),
            #                   int(fms[names.index(one_file_name)].split(')')[0].split('(')[1].split(',')[1].split(' ')[1]),
            #                   int(fms[names.index(one_file_name)].split(')')[0].split('(')[1].split(',')[1].split(' ')[2])))
            # print(fms[names.index(temp_name)].split(')')[0].split('(')[1].split(',')[0])
            # print(fms[names.index(temp_name)].split(')')[0].split('(')[1].split(',')[1].split(' ')[1])
        if not p_tls[names.index(one_file_name)]:
            # print("?")
            keypoints.append((0, 0))
        else:
            op_area = p_tls[names.index(one_file_name)].replace('(', '').replace(')', '').replace(' ', '').split(',')[
                0]  # op area
            tl_x = int(p_tls[names.index(one_file_name)].replace('(', '').replace(')', '').replace(' ', '').split(',')[
                           1])  # x coordinate
            tl_y = int(p_tls[names.index(one_file_name)].replace('(', '').replace(')', '').replace(' ', '').split(',')[
                           2])  # y coordinate
            if analysis_img.shape[1] < tl_x or analysis_img.shape[0] < tl_y:
                continue
            # cv2.line(temp_img, (tl_x, tl_y), (tl_x, tl_y), (0, 0, 255), 5)
            keypoints.append((op_area, (tl_x, tl_y)))
            # print(type(tls[names.index(one_file_name)]))
            # keypoints.append((int(tls[names.index(one_file_name)].split(')')[0].split('(')[1].split(',')[0]),
            #                   int(tls[names.index(one_file_name)].split(')')[0].split('(')[1].split(',')[1].split(' ')[1]),
            #                   int(tls[names.index(one_file_name)].split(')')[0].split('(')[1].split(',')[1].split(' ')[2])))
            # print(tls[names.index(temp_name)].split(')')[0].split('(')[1].split(',')[0])
            # print(tls[names.index(temp_name)].split(')')[0].split('(')[1].split(',')[1].split(' ')[1])
        if not p_tms[names.index(one_file_name)]:
            # print("?")
            keypoints.append((0, 0))
        else:
            op_area = p_tms[names.index(one_file_name)].replace('(', '').replace(')', '').replace(' ', '').split(',')[
                0]  # op area
            tm_x = int(p_tms[names.index(one_file_name)].replace('(', '').replace(')', '').replace(' ', '').split(',')[
                           1])  # x coordinate
            tm_y = int(p_tms[names.index(one_file_name)].replace('(', '').replace(')', '').replace(' ', '').split(',')[
                           2])  # y coordinate
            if analysis_img.shape[1] < tm_x or analysis_img.shape[0] < tm_y:
                continue
            # cv2.line(temp_img, (tm_x, tm_y), (tm_x, tm_y), (0, 0, 255), 5)
            keypoints.append((op_area, (tm_x, tm_y)))
            # print(type(tms[names.index(one_file_name)]))
            # keypoints.append((int(tms[names.index(one_file_name)].split(')')[0].split('(')[1].split(',')[0]),
            #                   int(tms[names.index(one_file_name)].split(')')[0].split('(')[1].split(',')[1].split(' ')[1]),
            #                   int(tms[names.index(one_file_name)].split(')')[0].split('(')[1].split(',')[1].split(' ')[2])))
            # print(tms[names.index(temp_name)].split(')')[0].split('(')[1].split(',')[0])
            # print(tms[names.index(temp_name)].split(')')[0].split('(')[1].split(',')[1].split(' ')[1])
    #     # print(keypoints)
    #     print(tempN)
    #     print(keypoints)
        # if '900782748R_joint.jpg' in tempN:
        print(keypoints)
        print(one_file_name)

        ## keypoints 찍는거 추가한 것
        val_img = cv2.imread('/app/pilot/validation/joint/' + one_file_name + '')
        #print(val_img)
        h, w, c = val_img.shape
        addy = (w - h) / 2

        for i in range(4):
            if keypoints[i][1] != 0:
                print(keypoints[0][1])
                print(keypoints[i][1][0],keypoints[i][1][1]+int(addy))
                cv2.line(temp_img, (keypoints[i][1][0],keypoints[i][1][1]+int(addy)), (keypoints[i][1][0],keypoints[i][1][1]+int(addy)), (255, 0, 0), 5)
                #cv2.line(temp_img, keypoints[i][1]+addy, keypoints[i][1]+addy, (255, 0, 0), 5)

#        cv2.line(temp_img, keypoints[0][1], keypoints[0][1], (255, 0, 0), 5)
#        cv2.line(temp_img, keypoints[1][1], keypoints[1][1], (255, 0, 0), 5)
#        cv2.line(temp_img, keypoints[2][1], keypoints[2][1], (255, 0, 0), 5)
#        cv2.line(temp_img, keypoints[3][1], keypoints[3][1], (255, 0, 0), 5)
        cv2.imwrite(os.path.join(save_path, 'test_' + one_file_name), temp_img)
#         cv2.imwrite('./oai_point_test' + '/' + temp_name, analysis_img)
        # if not os.path.isdir('./theshold_test/' + one_file_name.replace('_joint.jpg', '')):
        #     os.mkdir('./theshold_test/' + one_file_name.replace('_joint.jpg', ''))
        #     cv2.imwrite(os.path.join('./theshold_test', one_file_name.replace('_joint.jpg', ''), one_file_name), origin_img)
        # left_avg, left_std, right_avg, right_std, left_min, right_min = calculator_jsw_test_Threshold(left_avg, right_avg, left_std, right_std, analysis_img, keypoints, one_file_name, one_file_name, Threshold)
        # temp_list.append((one_file_name, kl, left_avg, left_std, right_avg, right_std, left_min, right_min))
        # add gs info

        left_avg, left_std, right_avg, right_std, left_min, right_min, gs_left_avg, gs_left_std, gs_right_avg, \
        gs_right_std, gs_left_min, gs_right_min, left_area, gs_left_area, right_area, gs_right_area = calculator_jsw_test_Threshold_compare_gs(left_avg, right_avg, left_std, right_std,
                                                                       analysis_img, keypoints, one_file_name, one_file_name, T)
        temp_list.append((one_file_name, kl, left_avg, left_std, right_avg, right_std, left_min, right_min, gs_left_avg,
                          gs_left_std, gs_right_avg, gs_right_std, gs_left_min, gs_right_min, left_area, gs_left_area, right_area, gs_right_area))
        # print("left avg : ", left_avg, 'left_std : ', left_std, 'right_avg : ', right_avg, 'right_std : ', right_std)
        #######     calculator_jsw_keypoint_test(analysis_img, keypoints, one_file_name)# original, keypoint_list, name

    save_to_excel_test_Threshold(temp_list, order, T)